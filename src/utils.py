import time
import json
import os
import logging
import re

from urllib.parse import parse_qs, urlparse
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import WebDriverException
from openpyxl import load_workbook

from navigation import navigate_to_login, perform_login, handle_checkpoint, wait_for_page_load
from vars import exp_folder, ext_folder

def initiate_browser():
    """Initializes the Chrome browser."""
    try:
        browser = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
        logging.info("New Chrome Driver Started")
        return browser
    except WebDriverException as e:
        logging.error(f"Error initializing Chrome browser: {e}")
        return None

def handle_login(browser, cookies_file):
    """Handles the login process and loads cookies if available."""
    browser.get("https://www.linkedin.com")
    if os.path.exists(cookies_file):
        load_cookies(browser, cookies_file)
        browser.refresh()
        wait_for_page_load(browser)
    else:
        navigate_to_login(browser)
        perform_login(browser)
        handle_checkpoint(browser)
        save_cookies(browser, cookies_file)

def save_cookies(browser, filepath):
    """Saves cookies to a file."""
    with open(filepath, 'w') as file:
        json.dump(browser.get_cookies(), file)
        logging.info("Saving cookies at " + filepath)

def load_cookies(browser, filepath):
    """Loads cookies from a file."""
    with open(filepath, 'r') as file:
        cookies = json.load(file)
        logging.info("Loading cookies from " + str(filepath))
        for cookie in cookies:
            # Add cookies to the browser
            if 'sameSite' in cookie:
                del cookie['sameSite']
            browser.add_cookie(cookie)

def set_job_url(default_url):
    domain = "linkedin.com"
    required_params = ["geoId", "keywords"]
    
    job_url = input("\nSet URL (Enter for default): ")
    if job_url:
        parsed_url = urlparse(job_url)
        if re.search(domain, parsed_url.netloc):
            query_params = parse_qs(parsed_url.query)
            if all(param in query_params for param in required_params):
                logging.info("URL updated")
                return job_url
        logging.warning("Invalid URL. Fallback to default")
        return default_url
    else:
        logging.info("Fallback to default")
        return default_url

def quit_browser(browser):
    """Quits the browser."""
    browser.quit()

def set_filename(url):
    t = time.localtime()
    current_time = time.strftime('%d-%m-%y_%H:%M:%S', t)
    filename = current_time
    parsed_url = urlparse(url)
    query_params = parse_qs(parsed_url.query)
    if 'keywords' in query_params:
        filename = filename + "-" + query_params['keywords'][0]
    else:
        filename = filename + "-" + "recommended"
    filename = filename.replace(' ', '-')
    filename = filename + '.txt'
    return filename

def file_save(filename, data):
    with open(ext_folder + filename, 'a') as file:
        for item in data:
            if item == "[]":
                continue
            file.write(f"{item}\n")

def set_excelname(url):
    filename = time.strftime('%d-%m-%y_%H:%M:%S', time.localtime())
    parsed_url = urlparse(url)
    query_params = parse_qs(parsed_url.query)
    if 'keywords' in query_params:
        filename = filename + "_" + query_params['keywords'][0]
        filename = filename + ".xlsx"
    else:
        filename = filename + "_" + url
        filename = filename.replace(ext_folder, '')
        filename = filename.replace('.txt','')
    filename = filename.replace(' ', '-')
    filename = exp_folder + filename + '.xlsx'
    return filename

def list_files_in_directory(directory):
    """Lists files in the given directory and returns a list of file paths."""
    files = os.listdir(directory)
    files = [os.path.join(directory, f) for f in files if os.path.isfile(os.path.join(directory, f))]
    return files

def save_to_excel(df, url, numjobs):
    file_name = set_excelname(url)
    df['Percentage'] = df['Percentage'] / 100

    # Save DataFrame to Excel
    df.to_excel(file_name, index=False, engine='openpyxl')

    # Load the workbook and select the active sheet
    wb = load_workbook(file_name)
    ws = wb.active
    
    ws.insert_rows(1)
    ws['A1'] = f"Number of jobs: {numjobs}"

    for cell in ws['A'][2:]:
        cell.number_format = '@'
    for cell in ws['B'][2:]:
        cell.number_format = '0'
    for cell in ws['C'][2:]:
        cell.number_format = '0%'
    
    # Save the workbook
    wb.save(file_name)